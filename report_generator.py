"""
report_generator.py — PDF, CSV, and Excel report generation.
"""

import io
import csv
from pathlib import Path
from datetime import datetime

REPORTS_DIR = Path("reports")
REPORTS_DIR.mkdir(exist_ok=True)

CLASSES = ["Organic", "Plastic", "Paper", "Mixed", "Concealed_Polybag"]


# ── CSV ───────────────────────────────────────────────────────────────────────

def generate_csv(run_id: int) -> bytes:
    from database import get_epoch_history, get_all_runs
    history = get_epoch_history(run_id)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=[
        "epoch", "train_loss", "val_loss", "train_acc",
        "val_acc", "learning_rate", "duration_seconds"
    ])
    writer.writeheader()
    writer.writerows(history)
    return buf.getvalue().encode()


def generate_all_runs_csv() -> bytes:
    from database import get_all_runs
    runs = get_all_runs()
    buf = io.StringIO()
    if runs:
        writer = csv.DictWriter(buf, fieldnames=runs[0].keys())
        writer.writeheader()
        writer.writerows(runs)
    return buf.getvalue().encode()


# ── Excel ─────────────────────────────────────────────────────────────────────

def generate_excel(run_id: int) -> bytes:
    from database import get_epoch_history, get_fl_rounds_for_run, get_all_runs
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="1a472a")
    header_font = Font(color="ffffff", bold=True)

    def style_header(ws, headers):
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 4)

    # Sheet 1 — Epoch Metrics
    ws1 = wb.active
    ws1.title = "Training Metrics"
    history = get_epoch_history(run_id)
    cols1 = ["Epoch", "Train Loss", "Val Loss", "Train Acc %",
             "Val Acc %", "Learning Rate", "Duration (s)"]
    style_header(ws1, cols1)
    for i, row in enumerate(history, 2):
        ws1.cell(i, 1, row["epoch"])
        ws1.cell(i, 2, round(row["train_loss"], 4))
        ws1.cell(i, 3, round(row["val_loss"], 4))
        ws1.cell(i, 4, round(row["train_acc"] * 100, 2))
        ws1.cell(i, 5, round(row["val_acc"] * 100, 2))
        ws1.cell(i, 6, row["learning_rate"])
        ws1.cell(i, 7, round(row["duration_seconds"], 2))

    # Sheet 2 — FL Rounds
    ws2 = wb.create_sheet("FL Rounds")
    fl_rounds = get_fl_rounds_for_run(run_id)
    cols2 = ["Round", "Node", "Local Loss", "Local Acc %",
             "CO2 (kg)", "Bandwidth (MB)", "Samples"]
    style_header(ws2, cols2)
    for i, row in enumerate(fl_rounds, 2):
        ws2.cell(i, 1, row["round_num"])
        ws2.cell(i, 2, row["node_id"])
        ws2.cell(i, 3, round(row["local_loss"], 4))
        ws2.cell(i, 4, round(row["local_acc"] * 100, 2))
        ws2.cell(i, 5, row["co2_kg"])
        ws2.cell(i, 6, row["bandwidth_mb"])
        ws2.cell(i, 7, row["num_samples"])

    # Sheet 3 — All Runs
    ws3 = wb.create_sheet("All Runs")
    all_runs = get_all_runs()
    cols3 = ["ID", "Mode", "Status", "Best Val Acc %",
             "Total Epochs", "Started At", "Finished At"]
    style_header(ws3, cols3)
    for i, r in enumerate(all_runs, 2):
        ws3.cell(i, 1, r["id"])
        ws3.cell(i, 2, r["mode"])
        ws3.cell(i, 3, r["status"])
        ws3.cell(i, 4, round((r["best_val_acc"] or 0) * 100, 2))
        ws3.cell(i, 5, r["total_epochs"])
        ws3.cell(i, 6, str(r["started_at"] or ""))
        ws3.cell(i, 7, str(r["finished_at"] or ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ───────────────────────────────────────────────────────────────────────

def generate_pdf(run_id: int, dataset_stats: dict = None) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable
    )
    from database import get_epoch_history, get_all_runs

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()

    GREEN = colors.HexColor("#22c55e")
    DARK = colors.HexColor("#0a0f0a")

    title_style = ParagraphStyle(
        "GreenTitle", parent=styles["Title"],
        textColor=GREEN, fontSize=22, spaceAfter=6
    )
    h2_style = ParagraphStyle(
        "GreenH2", parent=styles["Heading2"],
        textColor=GREEN, fontSize=13, spaceAfter=4
    )
    normal = styles["Normal"]

    history = get_epoch_history(run_id)
    runs = get_all_runs()
    this_run = next((r for r in runs if r["id"] == run_id), {})

    story = []

    # Title
    story.append(Paragraph("GreenScan — Training Report", title_style))
    story.append(Paragraph(
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | "
        f"Run #{run_id} | Mode: {this_run.get('mode', 'N/A').upper()}",
        normal
    ))
    story.append(HRFlowable(color=GREEN, thickness=1, spaceAfter=12))

    # Dataset stats
    if dataset_stats:
        story.append(Paragraph("Dataset Statistics", h2_style))
        ds_data = [["Split", "Total Images"] + CLASSES]
        for split, info in dataset_stats.get("splits", {}).items():
            row = [split.capitalize(), str(info["total"])]
            for cls in CLASSES:
                row.append(str(info["per_class"].get(cls, 0)))
            ds_data.append(row)
        ds_table = Table(ds_data, hAlign="LEFT")
        ds_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), DARK),
            ("TEXTCOLOR", (0, 0), (-1, 0), GREEN),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        story.append(ds_table)
        story.append(Spacer(1, 0.4*cm))

    # Training summary
    story.append(Paragraph("Training Summary", h2_style))
    summary_data = [
        ["Metric", "Value"],
        ["Status", this_run.get("status", "N/A")],
        ["Total Epochs", str(this_run.get("total_epochs", 0))],
        ["Best Val Accuracy", f"{round((this_run.get('best_val_acc') or 0)*100, 2)}%"],
        ["Started At", str(this_run.get("started_at", "N/A"))],
        ["Finished At", str(this_run.get("finished_at", "N/A"))],
    ]
    sum_table = Table(summary_data, hAlign="LEFT", colWidths=[8*cm, 8*cm])
    sum_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), GREEN),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
    ]))
    story.append(sum_table)
    story.append(Spacer(1, 0.4*cm))

    # Epoch table
    if history:
        story.append(Paragraph("Epoch-by-Epoch Metrics", h2_style))
        ep_data = [["Epoch", "Train Loss", "Val Loss", "Train Acc %", "Val Acc %", "LR"]]
        for row in history:
            ep_data.append([
                str(row["epoch"]),
                f"{row['train_loss']:.4f}",
                f"{row['val_loss']:.4f}",
                f"{row['train_acc']*100:.1f}",
                f"{row['val_acc']*100:.1f}",
                f"{row['learning_rate']:.2e}",
            ])
        ep_table = Table(ep_data, hAlign="LEFT")
        ep_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), DARK),
            ("TEXTCOLOR", (0, 0), (-1, 0), GREEN),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ]))
        story.append(ep_table)

    doc.build(story)
    return buf.getvalue()
